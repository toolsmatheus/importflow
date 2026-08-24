"""Gera data/reference/dcb-index.json a partir de data/reference/dcb.xlsx.

Lista consolidada Anvisa. Codigos com 5 digitos (ex: 1 vira 00001).

    python scripts/build_dcb_index.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Instale openpyxl: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "reference" / "dcb.xlsx"
OUT = ROOT / "data" / "reference" / "dcb-index.json"


def pad_dcb(value) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value).strip())
    if not digits:
        return None
    return digits.zfill(5)


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Arquivo não encontrado: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by_code: dict[str, str] = {}
    header_seen = False

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        first = str(row[0]).strip().upper()
        if not header_seen:
            if "DCB" in first or first.startswith("N"):
                header_seen = True
            continue
        code = pad_dcb(row[0])
        nome = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not code or not nome:
            continue
        by_code[code] = nome.upper()

    wb.close()
    payload = {
        "source": XLSX.name,
        "count": len(by_code),
        "byCode": by_code,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"OK {OUT} — {len(by_code)} DCBs")


if __name__ == "__main__":
    main()
