/**
 * Gera data/reference/cmed-ean-index.json a partir de data/reference/cmed.xlsx
 *
 * Uso (na raiz do repo, com Python + openpyxl):
 *   python scripts/build_cmed_index.py
 */
from __future__ import annotations

import json
import re
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Instale openpyxl: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "reference" / "cmed.xlsx"
OUT = ROOT / "data" / "reference" / "cmed-ean-index.json"


def clean_ean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.replace(" ", "").replace("-", "") in {"", "-"}:
        return None
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) >= 8 else None


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Arquivo não encontrado: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    header = None
    for row in rows:
        if row and any(c and "SUBST" in str(c).upper() for c in row):
            header = list(row)
            break
    if not header:
        raise SystemExit("Cabeçalho SUBSTÂNCIA não encontrado no XLSX")

    def col(*names: str) -> int | None:
        upper = [str(h).strip().upper() if h else "" for h in header]
        for name in names:
            if name.upper() in upper:
                return upper.index(name.upper())
        for name in names:
            for i, h in enumerate(upper):
                if name.upper() in h:
                    return i
        return None

    i_sub = col("SUBSTÂNCIA") or 1
    i_reg = col("REGISTRO") or 5
    i_e1 = col("EAN 1") or 6
    i_e2 = col("EAN 2") or 7
    i_e3 = col("EAN 3") or 8
    i_prod = col("PRODUTO") or 9
    i_tarja = col("TARJA")

    by_ean: dict[str, dict] = {}
    substances: set[str] = set()

    for row in rows:
        if not row or not row[i_sub]:
            continue
        subst = str(row[i_sub]).strip()
        substances.add(subst)
        meta = {
            "s": subst,
            "r": str(row[i_reg]).strip() if row[i_reg] else "",
            "p": str(row[i_prod]).strip() if row[i_prod] else "",
            "t": str(row[i_tarja]).strip() if i_tarja is not None and row[i_tarja] else "",
        }
        for ix in (i_e1, i_e2, i_e3):
            ean = clean_ean(row[ix] if ix is not None else None)
            if ean and ean not in by_ean:
                by_ean[ean] = meta

    wb.close()
    payload = {
        "source": XLSX.name,
        "eanCount": len(by_ean),
        "substanceCount": len(substances),
        "byEan": by_ean,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"OK {OUT} — {len(by_ean)} EANs, {len(substances)} substâncias")


if __name__ == "__main__":
    main()
